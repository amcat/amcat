###########################################################################
#          (C) Vrije Universiteit, Amsterdam (the Netherlands)            #
#                                                                         #
# This file is part of AmCAT - The Amsterdam Content Analysis Toolkit     #
#                                                                         #
# AmCAT is free software: you can redistribute it and/or modify it under  #
# the terms of the GNU Affero General Public License as published by the  #
# Free Software Foundation, either version 3 of the License, or (at your  #
# option) any later version.                                              #
#                                                                         #
# AmCAT is distributed in the hope that it will be useful, but WITHOUT    #
# ANY WARRANTY; without even the implied warranty of MERCHANTABILITY or   #
# FITNESS FOR A PARTICULAR PURPOSE. See the GNU Affero General Public     #
# License for more details.                                               #
#                                                                         #
# You should have received a copy of the GNU Affero General Public        #
# License along with AmCAT.  If not, see <http://www.gnu.org/licenses/>.  #
###########################################################################

"""
Plugin for uploading csv files
"""
import csv
import datetime
import itertools
import logging
import os
from collections import defaultdict
from operator import itemgetter
from typing import Optional, Tuple, Type, TypeVar, Callable, Any

import iso8601

from amcat.contrib.oset import OrderedSet
from amcat.models import Article, get_property_primitive_type
from amcat.scripts.article_upload.upload import ArticleField, UploadScript, _open
from amcat.scripts.article_upload.upload_plugins import UploadPlugin
from amcat.tools.amcates import ARTICLE_FIELDS
from amcat.tools.toolkit import read_date

log = logging.getLogger(__name__)

REQUIRED = tuple(
    field.name for field in Article._meta.get_fields() if field.name in ARTICLE_FIELDS and not field.blank)

PARSERS = {
    datetime.datetime: read_date,
    set: lambda s: {tag for tag in (part.strip() for part in s.split(",")) if tag},
}

FieldType = TypeVar("FieldType")


def get_parser(field_type: Type[FieldType]) -> Callable[[Any], FieldType]:
    return PARSERS.get(field_type, lambda x: field_type(x))


def parse_value(property: str, value: str):
    t = get_property_primitive_type(property)
    parser = get_parser(t)
    return parser(value)


def get_fieldname_tuple(field_name: str, suggested_type: str):
    if field_name.endswith("_" + suggested_type):
        return field_name, suggested_type
    return "{}_{}".format(field_name, suggested_type), suggested_type


def to_valid_field_name(field_name: str) -> str:
    """
    Filter field name so that it's a valid destination name.
    Valid names consist of an alphabetical character followed by any number of alphanumerical characters.
    @param field_name: A field name that can contain invalid characters
    @return: A valid destination name.
    """
    field_name = field_name[0] + field_name.title()[1:]
    if field_name[0].isdigit():
        field_name = "f" + field_name
    for char in field_name:
        if not char.isalnum():
            field_name = field_name.replace(char, "")
    return field_name


def guess_destination_and_type(field_name: str, sample_value: Optional[str]) -> Tuple:
    if field_name == "id":
        # The uploader *probably* does not mean to use the id field (possibly originating
        # from another AmCAT instance).
        return None, None

    from navigator.views.articleset_upload_views import CORE_FIELDS
    if field_name in CORE_FIELDS:
        return field_name, None

    # Guess based on field_name:
    ptype = get_property_primitive_type(field_name)
    if ptype != str:
        if ptype == float:
            return field_name, "num"
        elif ptype == int:
            return field_name, "int"
        elif ptype == datetime.datetime:
            return field_name, "date"
        else:
            raise ValueError("Did not recognize return value of get_property_primitive_type: {}".format(ptype))

    # Guess based on sample value
    if sample_value is None:
        # Not able to guess something, do not recommend anything
        return None, None

    # Guess datetime
    try:
        iso8601.parse_date(sample_value)
    except iso8601.ParseError:
        pass
    else:
        return get_fieldname_tuple(field_name, "date")

    # TODO: Guess datetime based on DATE_INPUT_TYPES of Django

    # Guess int
    try:
        int(sample_value)
    except ValueError:
        pass
    else:
        return get_fieldname_tuple(field_name, "int")

    # Guess float
    try:
        float(sample_value)
    except ValueError:
        pass
    else:
        return get_fieldname_tuple(field_name, "num")

    if sample_value.startswith(("https://", "http://")):
        return get_fieldname_tuple(field_name, "url")

    return field_name, "default"


class CsvDictReader(csv.DictReader):
    extensions = ["csv", "txt"]
    require_binary = False


class XlsxDictReader:
    """
    xlsx reader that is iterable and yields dicts.
    """
    extensions = ["xlsx"]
    require_binary = True

    def __init__(self, fp):
        from openpyxl import load_workbook
        self.fp = fp
        workbook = load_workbook(fp)
        self.worksheet = workbook.worksheets[0]
        self.fieldnames = [cell.value for cell in self.worksheet.rows[0]]

    def __iter__(self):
        rows = (x for i, x in enumerate(self.worksheet.iter_rows()) if i != 0)
        for row in rows:
            data = dict((k, str(v.value) if v.value is not None else "") for k, v in zip(self.fieldnames, row) if k)
            yield data


@UploadPlugin(label="CSV / XLSX", default=True,
              mime_types=("text/plain", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
class CSV(UploadScript):
    """
    Upload CSV files to AmCAT.

    To tell AmCAT which columns from the csv file to use, you need to specify the name in the file
    for the AmCAT-fields that you want to import. So, if you have a 'title' column in the csv file
    that you want to import as the headline, specify 'title' in the "headline field" input box.

    Text and date and required, all other fields are optional.

    If you are encountering difficulties, please make sure that you know how the csv is exported, and
    manually set encoding and dialect in the options above.

    Since Excel has quite some difficulties with exporting proper csv, it is often better to use
    an alternative such as OpenOffice or Google Spreadsheet (but see below for experimental xlsx support).
    If you must use excel, there is a 'tools' button on the save dialog which allows you to specify the
    encoding and delimiter used.

    We have added experimental support for .xlsx files (note: only use .xlsx, not the older .xls file type).
    This will hopefully alleviate some of the problems with reading Excel-generated csv file. Only the
    first sheet will be used, and please make sure that the data in that sheet has a header row. Please let
    us know if you encounter any difficulties at github.com/amcat/amcat/issues. Since you can only attach
    pictures there, the best way to share the file that you are having difficulty with (if it is not private)
    is to upload it to dropbox or a file sharing website and paste the link into the issue.
    """

    readers = {extension: reader for reader in (CsvDictReader, XlsxDictReader) for extension in reader.extensions}

    _errors = {
        "empty_col": 'Expected non-empty value in table column "{}" for required field "{}".',
        "empty_val": 'Expected non-empty value for required field "{}".',
        "parse_value": 'Failed to parse value "{}". Expected type: {}.'
    }

    def parse_value(self, property, value):
        t = get_property_primitive_type(property)
        parser = PARSERS[t]
        return parser(value)

    @classmethod
    def get_reader(cls, file, encoding):
        _, extension = os.path.splitext(file)
        reader = cls.readers[extension[1:]]
        if reader.require_binary:
            file = open(file, "rb")
        else:
            file = _open(file, encoding)
        return reader(file)

    def parse_file(self, file, encoding, _data):
        reader = self.get_reader(file, encoding)
        for unmapped_dict in reader:
            art_dict = self.map_article(unmapped_dict)
            properties = {}
            for k, v in art_dict.items():
                v = parse_value(k, v)
                properties[k] = v
            yield Article.fromdict(properties)

    def split_file(self, file):
        for reader in file:
            yield reader

    def explain_error(self, error, article=None):
        return "Error in row {}: {}".format(article, error)

    @classmethod
    def get_fields(cls, file, encoding):
        sample_data = defaultdict(OrderedSet)

        for f, enc, _ in UploadScript._get_files(file, encoding):
            reader = cls.get_reader(f, encoding)
            for row in itertools.islice(reader, 0, 5):
                for field_name, value in row.items():
                    if value.strip():
                        sample_data[field_name].add(value.strip())

        # Delete empty data
        for values in sample_data.values():
            if "" in values:
                values.remove("")

        # Guess types and destinations
        for field_name, values in sorted(sample_data.items(), key=itemgetter(0)):
            filtered_field_name = to_valid_field_name(field_name)
            try:
                value = next(iter(values))
            except StopIteration:
                value = None

            suggested_destination, suggested_type = guess_destination_and_type(filtered_field_name, value)
            yield ArticleField(field_name,
                               suggested_destination=suggested_destination,
                               values=list(itertools.islice(sample_data[field_name], 0, 5)),
                               suggested_type=suggested_type)


if __name__ == '__main__':
    from amcat.scripts.tools import cli

    cli.run_cli(CSV)
